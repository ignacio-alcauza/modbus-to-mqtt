from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Color palette ──────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", start_color="1F3864")   # dark navy
SUB_FILL   = PatternFill("solid", start_color="2E75B6")   # blue
GRP_FILL   = PatternFill("solid", start_color="D6E4F0")   # light blue
ROW_ALT    = PatternFill("solid", start_color="EBF3FB")   # very light blue
ROW_WHITE  = PatternFill("solid", start_color="FFFFFF")
RW_FILL    = PatternFill("solid", start_color="E2EFDA")   # green for R/W
RO_FILL    = PatternFill("solid", start_color="FFF2CC")   # yellow for R
ALARM_FILL = PatternFill("solid", start_color="FCE4D6")   # orange for alarms
TITLE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
GRP_FONT   = Font(name="Arial", bold=True, color="1F3864", size=10)
CELL_FONT  = Font(name="Arial", size=9)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="B0C4DE")
med  = Side(style="medium", color="2E75B6")
BORDER_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_MED  = Border(left=med,  right=med,  top=med,  bottom=med)

def style_header(cell, text, fill=HDR_FILL, font=HDR_FONT, align=CENTER):
    cell.value = text; cell.fill = fill; cell.font = font
    cell.alignment = align; cell.border = BORDER_THIN

def style_group(cell, text):
    cell.value = text; cell.fill = GRP_FILL; cell.font = GRP_FONT
    cell.alignment = LEFT; cell.border = BORDER_THIN

def style_cell(cell, value, align=CENTER, fill=ROW_WHITE):
    cell.value = value; cell.fill = fill; cell.font = CELL_FONT
    cell.alignment = align; cell.border = BORDER_THIN

# ═══════════════════════════════════════════════════════════════
# SHEET 1 — Datos en tiempo real (0x1200)
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "0x1200 Datos Tiempo Real"

# Title row
ws1.merge_cells("A1:H1")
tc = ws1["A1"]
tc.value = "JK BMS JK-PB2A16S20P — Bloque 0x1200 · Datos en Tiempo Real (Solo Lectura)"
tc.fill = PatternFill("solid", start_color="1F3864")
tc.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
tc.alignment = CENTER
ws1.row_dimensions[1].height = 28

# Sub-header
ws1.merge_cells("A2:H2")
sc = ws1["A2"]
sc.value = "Función Modbus: FC03 (Read Holding Registers)  |  Todos los registros son SOLO LECTURA"
sc.fill = PatternFill("solid", start_color="2E75B6")
sc.font = Font(name="Arial", italic=True, color="FFFFFF", size=9)
sc.alignment = CENTER

# Column headers
cols = ["Dirección HEX","Offset DEC","Tipo de dato","Bytes","R/W","Nombre / Tag","Descripción","Unidad / Escala"]
for i, c in enumerate(cols, 1):
    style_header(ws1.cell(3, i), c)
ws1.row_dimensions[3].height = 22

# ── Data rows ──────────────────────────────────────────────────
realtime_data = [
    # Section, addr_hex, offset, dtype, bytes, rw, tag, desc, unit
    ("TENSIONES DE CELDAS", None, None, None, None, None, None, None, None),
    (None,"0x1200",0,"UINT16",2,"R","CellVol1","Tensión celda 1","mV"),
    (None,"0x1202",2,"UINT16",2,"R","CellVol2","Tensión celda 2","mV"),
    (None,"0x1204",4,"UINT16",2,"R","CellVol3","Tensión celda 3","mV"),
    (None,"0x1206",6,"UINT16",2,"R","CellVol4","Tensión celda 4","mV"),
    (None,"0x1208",8,"UINT16",2,"R","CellVol5","Tensión celda 5","mV"),
    (None,"0x120A",10,"UINT16",2,"R","CellVol6","Tensión celda 6","mV"),
    (None,"0x120C",12,"UINT16",2,"R","CellVol7","Tensión celda 7","mV"),
    (None,"0x120E",14,"UINT16",2,"R","CellVol8","Tensión celda 8","mV"),
    (None,"0x1210",16,"UINT16",2,"R","CellVol9","Tensión celda 9","mV"),
    (None,"0x1212",18,"UINT16",2,"R","CellVol10","Tensión celda 10","mV"),
    (None,"0x1214",20,"UINT16",2,"R","CellVol11","Tensión celda 11","mV"),
    (None,"0x1216",22,"UINT16",2,"R","CellVol12","Tensión celda 12","mV"),
    (None,"0x1218",24,"UINT16",2,"R","CellVol13","Tensión celda 13","mV"),
    (None,"0x121A",26,"UINT16",2,"R","CellVol14","Tensión celda 14","mV"),
    (None,"0x121C",28,"UINT16",2,"R","CellVol15","Tensión celda 15","mV"),
    (None,"0x121E",30,"UINT16",2,"R","CellVol16","Tensión celda 16","mV"),
    ("RESISTENCIAS INTERNAS DE CELDAS", None, None, None, None, None, None, None, None),
    (None,"0x1220",32,"UINT16",2,"R","CellRes1","Resistencia interna celda 1","0.001 mΩ"),
    (None,"0x1222",34,"UINT16",2,"R","CellRes2","Resistencia interna celda 2","0.001 mΩ"),
    (None,"0x1224",36,"UINT16",2,"R","CellRes3","Resistencia interna celda 3","0.001 mΩ"),
    (None,"0x1226",38,"UINT16",2,"R","CellRes4","Resistencia interna celda 4","0.001 mΩ"),
    (None,"0x1228",40,"UINT16",2,"R","CellRes5","Resistencia interna celda 5","0.001 mΩ"),
    (None,"0x122A",42,"UINT16",2,"R","CellRes6","Resistencia interna celda 6","0.001 mΩ"),
    (None,"0x122C",44,"UINT16",2,"R","CellRes7","Resistencia interna celda 7","0.001 mΩ"),
    (None,"0x122E",46,"UINT16",2,"R","CellRes8","Resistencia interna celda 8","0.001 mΩ"),
    (None,"0x1230",48,"UINT16",2,"R","CellRes9","Resistencia interna celda 9","0.001 mΩ"),
    (None,"0x1232",50,"UINT16",2,"R","CellRes10","Resistencia interna celda 10","0.001 mΩ"),
    (None,"0x1234",52,"UINT16",2,"R","CellRes11","Resistencia interna celda 11","0.001 mΩ"),
    (None,"0x1236",54,"UINT16",2,"R","CellRes12","Resistencia interna celda 12","0.001 mΩ"),
    (None,"0x1238",56,"UINT16",2,"R","CellRes13","Resistencia interna celda 13","0.001 mΩ"),
    (None,"0x123A",58,"UINT16",2,"R","CellRes14","Resistencia interna celda 14","0.001 mΩ"),
    (None,"0x123C",60,"UINT16",2,"R","CellRes15","Resistencia interna celda 15","0.001 mΩ"),
    (None,"0x123E",62,"UINT16",2,"R","CellRes16","Resistencia interna celda 16","0.001 mΩ"),
    ("ESTADÍSTICAS DE TENSIÓN DEL PACK", None, None, None, None, None, None, None, None),
    (None,"0x1240",64,"UINT32",4,"R","VolCellAvg","Tensión media de celdas","mV"),
    (None,"0x1244",68,"UINT16",2,"R","VolCellMax","Tensión máxima de celda","mV"),
    (None,"0x1246",70,"UINT8",1,"R","NumCellMax","Número de celda con tensión máxima","—"),
    (None,"0x1248",72,"UINT16",2,"R","VolCellMin","Tensión mínima de celda","mV"),
    (None,"0x124A",74,"UINT8",1,"R","NumCellMin","Número de celda con tensión mínima","—"),
    (None,"0x124C",76,"UINT32",4,"R","VolCellDif","Diferencia max−min de tensión entre celdas","mV"),
    ("TEMPERATURAS", None, None, None, None, None, None, None, None),
    (None,"0x1260",96,"INT16",2,"R","TempBat1","Temperatura sensor batería 1","0.1 °C"),
    (None,"0x1262",98,"INT16",2,"R","TempBat2","Temperatura sensor batería 2","0.1 °C"),
    (None,"0x1264",100,"INT16",2,"R","TempMos","Temperatura MOSFET","0.1 °C"),
    (None,"0x12F8",248,"INT16",2,"R","TempBat3","Temperatura sensor batería 3","0.1 °C"),
    (None,"0x12FA",250,"INT16",2,"R","TempBat4","Temperatura sensor batería 4","0.1 °C"),
    ("CORRIENTE Y TENSIÓN DEL PACK", None, None, None, None, None, None, None, None),
    (None,"0x12E4",228,"UINT16",2,"R","BatVol","Tensión total del pack","0.01 V"),
    (None,"0x1280",128,"INT32",4,"R","CurBat","Corriente del pack (+ carga, − descarga)","mA"),
    (None,"0x12E6",230,"INT16",2,"R","HeatCurrent","Corriente de calefacción","mA"),
    ("SOC Y CAPACIDAD", None, None, None, None, None, None, None, None),
    (None,"0x12A4",164,"UINT32",4,"R","CapBatRem","Capacidad restante","mAh"),
    (None,"0x12A6",166,"UINT8",1,"R","SOC","Estado de carga (State of Charge)","% (0–100)"),
    (None,"0x12A7",167,"UINT8",1,"R","BalanStatus","Estado del balanceador (0=OFF, 1=carga, 2=descarga)","—"),
    (None,"0x12A8",168,"UINT32",4,"R","CapBatCyc","Capacidad acumulada en ciclos","mAh"),
    (None,"0x12AC",172,"UINT32",4,"R","CycCount","Contador de ciclos completos","—"),
    ("ESTADO Y ALARMAS", None, None, None, None, None, None, None, None),
    (None,"0x12B0",176,"UINT32",4,"R","AlarmState","Flags de alarma activas (ver hoja Alarmas)","bits"),
    (None,"0x12B4",180,"UINT32",4,"R","BalanceBit","Flags de balanceo activo por celda","bits"),
    (None,"0x12B8",184,"UINT32",4,"R","SysStatus","Estado del sistema (ver hoja Estado)","bits"),
    ("OTROS PARÁMETROS", None, None, None, None, None, None, None, None),
    (None,"0x12D4",212,"UINT16",2,"R","TimeEmergency","Tiempo en modo emergencia","s"),
    (None,"0x12D6",214,"UINT16",2,"R","BatCurCorrect","Factor corrección de corriente","—"),
    (None,"0x12D8",216,"UINT16",2,"R","VolChargCur","Tensión de referencia corriente de carga","mV"),
    (None,"0x12DA",218,"UINT16",2,"R","VolDischargCur","Tensión de referencia corriente de descarga","mV"),
    (None,"0x12DC",220,"FLOAT",4,"R","BatVolCorrect","Factor corrección de tensión pack","—"),
    (None,"0x12F0",240,"UINT32",4,"R","SysRunTicks","Tiempo de funcionamiento acumulado","0.1 s"),
]

row = 4
for d in realtime_data:
    if d[0] is not None and d[1] is None:
        ws1.merge_cells(f"A{row}:H{row}")
        style_group(ws1.cell(row, 1), f"▶  {d[0]}")
        ws1.row_dimensions[row].height = 18
        row += 1
        continue
    fill = ROW_ALT if (row % 2 == 0) else ROW_WHITE
    for col, val in enumerate([d[1],d[2],d[3],d[4],d[5],d[6],d[7],d[8]], 1):
        cell = ws1.cell(row, col)
        if col == 6:
            style_cell(cell, val, LEFT, fill)
        elif col == 7:
            style_cell(cell, val, LEFT, fill)
        else:
            style_cell(cell, val, CENTER, fill)
        if col == 5:
            cell.fill = RO_FILL
    row += 1

# Column widths
ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 12
ws1.column_dimensions["C"].width = 11
ws1.column_dimensions["D"].width = 8
ws1.column_dimensions["E"].width = 7
ws1.column_dimensions["F"].width = 18
ws1.column_dimensions["G"].width = 38
ws1.column_dimensions["H"].width = 16
ws1.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════
# SHEET 2 — Parámetros configuración (0x1000)
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("0x1000 Configuración")

ws2.merge_cells("A1:H1")
tc2 = ws2["A1"]
tc2.value = "JK BMS JK-PB2A16S20P — Bloque 0x1000 · Parámetros de Configuración (Lectura / Escritura)"
tc2.fill = PatternFill("solid", start_color="1F3864")
tc2.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
tc2.alignment = CENTER
ws2.row_dimensions[1].height = 28

ws2.merge_cells("A2:H2")
sc2 = ws2["A2"]
sc2.value = "Función Modbus: FC03 (Read) · FC10 (Write Multiple Registers)  |  ⚠ Modificar solo si conoces el efecto del parámetro"
sc2.fill = PatternFill("solid", start_color="C00000")
sc2.font = Font(name="Arial", italic=True, color="FFFFFF", size=9)
sc2.alignment = CENTER

for i, c in enumerate(cols, 1):
    style_header(ws2.cell(3, i), c)
ws2.row_dimensions[3].height = 22

config_data = [
    ("PROTECCIONES DE TENSIÓN", None, None, None, None, None, None, None, None),
    (None,"0x1000",0,"UINT32",4,"RW","VolSmartSleep","Tensión de entrada en modo sleep inteligente","mV"),
    (None,"0x1004",4,"UINT32",4,"RW","VolCellUV","Protección subtensión de celda","mV"),
    (None,"0x1008",8,"UINT32",4,"RW","VolCellUVPR","Recuperación de subtensión de celda","mV"),
    (None,"0x100C",12,"UINT32",4,"RW","VolCellOV","Protección sobretensión de celda","mV"),
    (None,"0x1010",16,"UINT32",4,"RW","VolCellOVPR","Recuperación de sobretensión de celda","mV"),
    (None,"0x1028",40,"UINT32",4,"RW","VolSysPwrOff","Tensión de apagado automático del sistema","mV"),
    (None,"0x105C",92,"UINT32",4,"RW","VolSysPwrOn","Tensión de encendido automático del sistema","mV"),
    ("BALANCEO", None, None, None, None, None, None, None, None),
    (None,"0x1014",20,"UINT32",4,"RW","VolBalanTrig","Diferencia de tensión para activar balanceo","mV"),
    ("REFERENCIA SOC", None, None, None, None, None, None, None, None),
    (None,"0x1018",24,"UINT32",4,"RW","VolSOC100","Tensión de celda correspondiente a SOC 100%","mV"),
    (None,"0x101C",28,"UINT32",4,"RW","VolSOC0","Tensión de celda correspondiente a SOC 0%","mV"),
    ("VOLTAJES DE CARGA", None, None, None, None, None, None, None, None),
    (None,"0x1020",32,"UINT32",4,"RW","VolCellRCV","Tensión de carga recomendada por celda","mV"),
    (None,"0x1024",36,"UINT32",4,"RW","VolCellRFV","Tensión de carga flotante por celda","mV"),
    ("PROTECCIONES DE CORRIENTE", None, None, None, None, None, None, None, None),
    (None,"0x102C",44,"UINT32",4,"RW","CurBatCOC","Corriente máxima de carga continua","mA"),
    (None,"0x1030",48,"UINT32",4,"RW","CurBatDOC","Corriente máxima de descarga continua","mA"),
    (None,"0x1034",52,"UINT32",4,"RW","CurBatSCP","Corriente de protección cortocircuito","mA"),
    ("TIEMPOS DE PROTECCIÓN", None, None, None, None, None, None, None, None),
    (None,"0x1038",56,"UINT32",4,"RW","TimBatSCP","Retardo protección cortocircuito","µs"),
    (None,"0x103C",60,"UINT32",4,"RW","TimBatCOP","Retardo protección sobrecorriente de carga","ms"),
    (None,"0x1040",64,"UINT32",4,"RW","TimBatDOP","Retardo protección sobrecorriente de descarga","ms"),
    ("PROTECCIONES DE TEMPERATURA", None, None, None, None, None, None, None, None),
    (None,"0x1044",68,"UINT32",4,"RW","TempBatCOT","Protección sobretemperatura de carga","0.1 °C"),
    (None,"0x1048",72,"UINT32",4,"RW","TempBatCOTPR","Recuperación sobretemperatura de carga","0.1 °C"),
    (None,"0x104C",76,"UINT32",4,"RW","TempBatDOT","Protección sobretemperatura de descarga","0.1 °C"),
    (None,"0x1050",80,"UINT32",4,"RW","TempBatDOTPR","Recuperación sobretemperatura de descarga","0.1 °C"),
    (None,"0x1054",84,"UINT32",4,"RW","TempBatCUT","Protección baja temperatura de carga","0.1 °C"),
    (None,"0x1058",88,"UINT32",4,"RW","TempBatCUTPR","Recuperación baja temperatura de carga","0.1 °C"),
    ("CAPACIDAD", None, None, None, None, None, None, None, None),
    (None,"0x1060",96,"UINT32",4,"RW","CapBat","Capacidad nominal de la batería","mAh"),
    (None,"0x1064",100,"UINT32",4,"RW","CapBatRem","Capacidad restante (calibración SOC)","mAh"),
    (None,"0x1068",104,"UINT32",4,"RW","CapBatCyc","Capacidad acumulada en ciclos","mAh"),
    ("CONFIGURACIÓN GENERAL", None, None, None, None, None, None, None, None),
    (None,"0x106C",108,"UINT32",4,"RW","CellCount","Número de celdas en serie","—"),
    ("INTERRUPTORES DE CONTROL", None, None, None, None, None, None, None, None),
    (None,"0x1070",112,"UINT32",4,"RW","SwitchCharge","Interruptor de carga (0=OFF, 1=ON)","bool"),
    (None,"0x1074",116,"UINT32",4,"RW","SwitchDischarge","Interruptor de descarga (0=OFF, 1=ON)","bool"),
    (None,"0x1078",120,"UINT32",4,"RW","SwitchBalance","Interruptor de balanceo (0=OFF, 1=ON)","bool"),
    (None,"0x107C",124,"UINT32",4,"RW","SwitchBuzzer","Interruptor zumbador (0=OFF, 1=ON)","bool"),
]

row = 4
for d in config_data:
    if d[0] is not None and d[1] is None:
        ws2.merge_cells(f"A{row}:H{row}")
        style_group(ws2.cell(row, 1), f"▶  {d[0]}")
        ws2.row_dimensions[row].height = 18
        row += 1
        continue
    fill = ROW_ALT if (row % 2 == 0) else ROW_WHITE
    for col, val in enumerate([d[1],d[2],d[3],d[4],d[5],d[6],d[7],d[8]], 1):
        cell = ws2.cell(row, col)
        if col in (6,7):
            style_cell(cell, val, LEFT, fill)
        else:
            style_cell(cell, val, CENTER, fill)
        if col == 5:
            cell.fill = RW_FILL
    row += 1

for col in ["A","B","C","D","E","F","G","H"]:
    ws2.column_dimensions[col].width = ws1.column_dimensions[col].width
ws2.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════
# SHEET 3 — Alarmas y bits de estado
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Alarmas y Estado")

ws3.merge_cells("A1:F1")
t3 = ws3["A1"]
t3.value = "JK BMS — Decodificación de Flags: AlarmState (0x12B0) · BalanceBit (0x12B4) · SysStatus (0x12B8)"
t3.fill = PatternFill("solid", start_color="1F3864")
t3.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
t3.alignment = CENTER
ws3.row_dimensions[1].height = 26

def alarm_section(ws, start_row, section_title, reg_addr, bit_data, fill_color):
    ws.merge_cells(f"A{start_row}:F{start_row}")
    sh = ws.cell(start_row, 1)
    sh.value = f"  {section_title}  —  Registro: {reg_addr}  (UINT32, 4 bytes)"
    sh.fill = PatternFill("solid", start_color=fill_color)
    sh.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    sh.alignment = LEFT
    ws.row_dimensions[start_row].height = 20
    start_row += 1

    hdrs = ["Bit","Máscara HEX","Nombre","Descripción","Acción sugerida","Gravedad"]
    for i, h in enumerate(hdrs, 1):
        style_header(ws.cell(start_row, i), h,
                     fill=PatternFill("solid", start_color="2E75B6"))
    ws.row_dimensions[start_row].height = 18
    start_row += 1

    for bit, mask, name, desc, action, severity in bit_data:
        fill = ALARM_FILL if severity in ("Alta","Crítica") else (ROW_ALT if (start_row%2==0) else ROW_WHITE)
        vals = [bit, mask, name, desc, action, severity]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(start_row, col)
            style_cell(cell, val, CENTER if col < 4 else LEFT, fill)
        start_row += 1
    return start_row + 1

alarm_bits = [
    (0,"0x00000001","OV_Cell","Sobretensión de celda","Verificar cargador; reducir tensión de carga","Alta"),
    (1,"0x00000002","UV_Cell","Subtensión de celda","Cargar batería de inmediato","Alta"),
    (2,"0x00000004","OC_Charge","Sobrecorriente de carga","Reducir corriente del cargador","Alta"),
    (3,"0x00000008","OC_Discharge","Sobrecorriente de descarga","Reducir carga conectada","Alta"),
    (4,"0x00000010","OT_Charge","Sobretemperatura en carga","Mejorar ventilación; reducir corriente","Media"),
    (5,"0x00000020","OT_Discharge","Sobretemperatura en descarga","Reducir carga; mejorar refrigeración","Media"),
    (6,"0x00000040","UT_Charge","Baja temperatura en carga","No cargar por debajo de 0 °C","Media"),
    (7,"0x00000080","SCP","Cortocircuito detectado","Verificar cableado y cargas","Crítica"),
    (8,"0x00000100","MOSFET_Charge_Fail","Fallo MOSFET de carga","Inspección técnica necesaria","Alta"),
    (9,"0x00000200","MOSFET_Discharge_Fail","Fallo MOSFET de descarga","Inspección técnica necesaria","Alta"),
    (10,"0x00000400","Temp_Sensor_Fail","Fallo sensor temperatura","Verificar conexión sensores NTC","Media"),
    (11,"0x00000800","Cell_Imbalance","Desequilibrio severo entre celdas","Ciclo de carga/descarga completo","Baja"),
]

balance_bits = [
    (0,"0x00000001","Bal_Cell1","Balanceo activo en celda 1","—","Info"),
    (1,"0x00000002","Bal_Cell2","Balanceo activo en celda 2","—","Info"),
    (2,"0x00000004","Bal_Cell3","Balanceo activo en celda 3","—","Info"),
    (3,"0x00000008","Bal_Cell4","Balanceo activo en celda 4","—","Info"),
    (4,"0x00000010","Bal_Cell5","Balanceo activo en celda 5","—","Info"),
    (5,"0x00000020","Bal_Cell6","Balanceo activo en celda 6","—","Info"),
    (6,"0x00000040","Bal_Cell7","Balanceo activo en celda 7","—","Info"),
    (7,"0x00000080","Bal_Cell8","Balanceo activo en celda 8","—","Info"),
    (8,"0x00000100","Bal_Cell9","Balanceo activo en celda 9","—","Info"),
    (9,"0x00000200","Bal_Cell10","Balanceo activo en celda 10","—","Info"),
    (10,"0x00000400","Bal_Cell11","Balanceo activo en celda 11","—","Info"),
    (11,"0x00000800","Bal_Cell12","Balanceo activo en celda 12","—","Info"),
    (12,"0x00001000","Bal_Cell13","Balanceo activo en celda 13","—","Info"),
    (13,"0x00002000","Bal_Cell14","Balanceo activo en celda 14","—","Info"),
    (14,"0x00004000","Bal_Cell15","Balanceo activo en celda 15","—","Info"),
    (15,"0x00008000","Bal_Cell16","Balanceo activo en celda 16","—","Info"),
]

status_bits = [
    (0,"0x00000001","Charging","MOS de carga cerrado (cargando)","—","Info"),
    (1,"0x00000002","Discharging","MOS de descarga cerrado (descargando)","—","Info"),
    (2,"0x00000004","Balancing","Balanceo en curso","—","Info"),
    (3,"0x00000008","Charging_Float","Carga flotante activa","—","Info"),
    (4,"0x00000010","Heating","Sistema de calefacción activo","—","Info"),
    (5,"0x00000020","Standby","BMS en modo standby","—","Info"),
    (6,"0x00000040","Full_Charge","Batería completamente cargada","—","Info"),
    (7,"0x00000080","Alarm_Active","Hay al menos una alarma activa","Ver registro AlarmState","Alta"),
]

r = alarm_section(ws3, 2, "ALARMAS (AlarmState)", "0x12B0", alarm_bits, "C00000")
r = alarm_section(ws3, r, "BALANCEO POR CELDA (BalanceBit)", "0x12B4", balance_bits, "375623")
r = alarm_section(ws3, r, "ESTADO DEL SISTEMA (SysStatus)", "0x12B8", status_bits, "2E75B6")

ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 22
ws3.column_dimensions["D"].width = 36
ws3.column_dimensions["E"].width = 34
ws3.column_dimensions["F"].width = 12
ws3.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════
# SHEET 4 — Guía de integración EW11A
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Guía EW11A")

ws4.merge_cells("A1:D1")
t4 = ws4["A1"]
t4.value = "Guía de Integración — Elfin EW11A + JK BMS JK-PB2A16S20P vía Modbus TCP"
t4.fill = PatternFill("solid", start_color="1F3864")
t4.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
t4.alignment = CENTER
ws4.row_dimensions[1].height = 28

guide = [
    ("PARÁMETROS DE COMUNICACIÓN RS485","","",""),
    ("Parámetro","Valor","Notas",""),
    ("Baudrate","115200 bps","Fijo, no modificable",""),
    ("Bits de datos","8","—",""),
    ("Paridad","None","—",""),
    ("Bits de parada","1","8N1",""),
    ("Protocolo BMS","JK BMS RS485 Modbus V1.1","Seleccionar en menú BMS: UART1 → 001",""),
    ("","","",""),
    ("CONFIGURACIÓN EW11A","","",""),
    ("Parámetro","Valor","Notas",""),
    ("Modo de trabajo","TCP Server","Puerto por defecto: 8899",""),
    ("Baudrate UART","115200","Debe coincidir con BMS",""),
    ("Paridad UART","None","8N1",""),
    ("IP","Asignada por DHCP o estática","Configurar IP fija para fiabilidad",""),
    ("Puerto TCP","8899","Puerto por defecto Elfin EW11A",""),
    ("","","",""),
    ("DIRECCIÓN ESCLAVO MODBUS (DIP Switch)","","",""),
    ("DIP 1","DIP 2","DIP 3","Dirección Modbus"),
    ("OFF","OFF","ON","1"),
    ("ON","OFF","OFF","2"),  # simplified
    ("OFF","ON","OFF","4 (ejemplo)",""),
    ("Nota","Todos OFF = inválido (no usar)","Usar valores 1–15",""),
    ("","","",""),
    ("LECTURA DE EJEMPLO — Tensión del pack (Python)","","",""),
    ("","","",""),
    ("# pip install pymodbus","","",""),
    ("from pymodbus.client import ModbusTcpClient","","",""),
    ("client = ModbusTcpClient('192.168.1.100', port=8899)","","",""),
    ("client.connect()","","",""),
    ("# Leer BatVol (0x12E4) → 1 registro","","",""),
    ("result = client.read_holding_registers(0x12E4, count=1, slave=1)","","",""),
    ("voltage = result.registers[0] * 0.01  # → Voltios","","",""),
    ("print(f'Tensión pack: {voltage:.2f} V')","","",""),
    ("client.close()","","",""),
    ("","","",""),
    ("NOTAS IMPORTANTES","","",""),
    ("1","Conectar al puerto RS485 derecho de la BMS (no al izquierdo, que es para inversor/CAN)","",""),
    ("2","Los registros UINT32 ocupan 2 registros Modbus consecutivos (big-endian)","",""),
    ("3","La corriente CurBat es INT32 (signed): positivo = carga, negativo = descarga","",""),
    ("4","Las temperaturas son INT16 en unidades de 0.1 °C → dividir entre 10","",""),
    ("5","Máximo recomendado de registros por petición FC03: 64 registros (128 bytes)","",""),
]

row = 2
for g in guide:
    is_section = g[1] == "" and g[2] == "" and g[3] == "" and g[0] and not g[0].startswith("#") and not g[0].startswith("from") and not g[0].startswith("client") and not g[0].startswith("result") and not g[0].startswith("voltage") and not g[0].startswith("print") and not g[0].startswith("1") and not g[0].startswith("2") and not g[0].startswith("3") and not g[0].startswith("4") and not g[0].startswith("5")
    is_header = g[1] == "Valor"

    if is_section:
        ws4.merge_cells(f"A{row}:D{row}")
        c = ws4.cell(row, 1)
        c.value = g[0]; c.fill = SUB_FILL
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.alignment = LEFT; c.border = BORDER_THIN
    elif is_header:
        for col, val in enumerate(g, 1):
            style_header(ws4.cell(row, col), val)
    elif g[0] == "":
        pass
    else:
        fill = ROW_ALT if (row % 2 == 0) else ROW_WHITE
        for col, val in enumerate(g, 1):
            c = ws4.cell(row, col)
            c.value = val; c.fill = fill; c.font = Font(name="Courier New" if col==1 and g[0].startswith(("from","client","result","voltage","print","#")) else "Arial", size=9)
            c.alignment = LEFT; c.border = BORDER_THIN
    ws4.row_dimensions[row].height = 16
    row += 1

ws4.column_dimensions["A"].width = 50
ws4.column_dimensions["B"].width = 30
ws4.column_dimensions["C"].width = 30
ws4.column_dimensions["D"].width = 20

# Save
output = "./JK_BMS_PB2A16S20P_Modbus_Map.xlsx"
wb.save(output)
print("Saved:", output)
